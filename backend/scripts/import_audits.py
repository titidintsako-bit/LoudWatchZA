"""
Import AGSA 2023-24 audit outcomes into Supabase municipalities table.
Reads from: data/raw/Governance and Financials/Annexure 1 - Auditees' audit outcomes 2023-24.xlsx

Sheet: "Annexure 1"
Data starts at row 8 (1-indexed), header at row 7.
Columns (0-indexed):
  0 = Number
  1 = Auditee name
  2 = Province (2-letter code: WC, EC, LP, etc.)
  3 = Auditee category
  4 = Auditee type
  5 = Municipal district
  6 = 2023-24 audit opinion

Usage:
    python scripts/import_audits.py
"""

import os
import logging
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).resolve().parent.parent / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

DATA_FILE = (
    Path(__file__).resolve().parent.parent.parent
    / "data" / "raw" / "Governance and Financials"
    / "Annexure 1 - Auditees' audit outcomes 2023-24.xlsx"
)

# Map 2-letter province codes to full names used in DB
PROVINCE_MAP = {
    "WC": "Western Cape",
    "EC": "Eastern Cape",
    "NC": "Northern Cape",
    "FS": "Free State",
    "KZN": "KwaZulu-Natal",
    "NW": "North West",
    "GP": "Gauteng",
    "MP": "Mpumalanga",
    "LP": "Limpopo",
    # Variants
    "KN": "KwaZulu-Natal",
    "LIM": "Limpopo",
    "GAU": "Gauteng",
}

# Map audit opinion text to numeric score (higher = better governance)
OPINION_SCORE = {
    "Unqualified – no findings": 5,
    "Unqualified – with findings": 4,
    "Qualified – with findings": 3,
    "Adverse – with findings": 2,
    "Disclaimed – with findings": 1,
    "Outstanding": 0,
    "Not submitted": 0,
}


def opinion_to_score(opinion: str | None) -> int:
    if not opinion:
        return 0
    opinion = str(opinion).strip()
    for key, score in OPINION_SCORE.items():
        if key.lower() in opinion.lower():
            return score
    return 0


def load_audits() -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(DATA_FILE, read_only=True, data_only=True)

        # Find the correct sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "annexure 1" in name.lower() or "annex" in name.lower():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        records = []
        for row in ws.iter_rows(min_row=8, values_only=True):
            auditee = row[1] if len(row) > 1 else None
            province_code = row[2] if len(row) > 2 else None
            opinion = row[6] if len(row) > 6 else None

            if not auditee or not province_code:
                continue

            province_code = str(province_code).strip().upper()
            province_name = PROVINCE_MAP.get(province_code, "")

            records.append({
                "name": str(auditee).strip(),
                "province": province_name or province_code,
                "province_code": province_code,
                "audit_opinion": str(opinion).strip() if opinion else "Unknown",
                "audit_score": opinion_to_score(opinion),
            })

        wb.close()
        logger.info(f"Loaded {len(records)} audit records from Excel.")
        return records

    except Exception as e:
        logger.error(f"Failed to read audit Excel: {e}")
        return []


def _get_supabase_client():
    from supabase import create_client
    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_KEY", "") or os.environ.get("SUPABASE_KEY", "")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_KEY must be set.")
    return create_client(url, key)


def run() -> None:
    records = load_audits()
    if not records:
        logger.error("No audit records loaded. Aborting.")
        return

    supabase = _get_supabase_client()

    # Fetch existing municipalities to match by name
    logger.info("Fetching existing municipalities from DB …")
    existing = supabase.table("municipalities").select("id,name,province").execute()
    muni_map: dict[str, str] = {}
    for m in (existing.data or []):
        key = str(m["name"]).strip().lower()
        muni_map[key] = m["id"]

    # Try to match audit records to municipalities
    matched = 0
    unmatched = []

    for rec in records:
        key = rec["name"].lower()
        muni_id = muni_map.get(key)

        if muni_id:
            supabase.table("municipalities").update({
                "audit_outcome": rec["audit_opinion"],
                "audit_score": rec["audit_score"],
            }).eq("id", muni_id).execute()
            matched += 1
        else:
            unmatched.append(rec["name"])

    logger.info(f"Matched and updated: {matched} municipalities")
    if unmatched:
        logger.info(f"Unmatched ({len(unmatched)}): {unmatched[:10]}{'...' if len(unmatched) > 10 else ''}")

    # Summary by opinion type
    from collections import Counter
    opinion_counts = Counter(r["audit_opinion"] for r in records)
    print(f"\nSummary:")
    print(f"  Total audit records:   {len(records)}")
    print(f"  Matched to DB:         {matched}")
    print(f"  Unmatched:             {len(unmatched)}")
    print(f"\nOpinion breakdown:")
    for opinion, count in sorted(opinion_counts.items(), key=lambda x: -x[1]):
        print(f"  {opinion:40s}  {count}")


if __name__ == "__main__":
    run()
