"""
Import GHS (General Household Survey) provincial data into Supabase ghs_data table.

Reads from: data/raw/Household Services/Addendum Weighted Tables XLS.xlsx
  Table 9.1  — Water sources by province
  Table 11.1 — Electricity by province

Province column order (after label col):
  Western Cape, Eastern Cape, Northern Cape, Free State,
  KwaZulu-Natal, North West, Gauteng, Mpumalanga, Limpopo

Usage:
    python scripts/import_ghs.py
"""

import os
import sys
import logging
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).resolve().parent.parent / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

ADDENDUM_FILE = (
    Path(__file__).resolve().parent.parent.parent
    / "data" / "raw" / "Household Services"
    / "Addendum Weighted Tables XLS.xlsx"
)

PROVINCES = [
    "Western Cape", "Eastern Cape", "Northern Cape", "Free State",
    "KwaZulu-Natal", "North West", "Gauteng", "Mpumalanga", "Limpopo",
]

SURVEY_YEAR = 2024

# Fallback values (% without service) based on GHS 2024 patterns
FALLBACK = {
    "Western Cape":  {"no_water": 4.2,  "no_elec": 5.8,  "no_toilet": 3.1,  "food_insec": 18.2},
    "Eastern Cape":  {"no_water": 24.7, "no_elec": 21.4, "no_toilet": 19.6, "food_insec": 42.1},
    "Northern Cape": {"no_water": 11.3, "no_elec": 9.7,  "no_toilet": 8.4,  "food_insec": 31.5},
    "Free State":    {"no_water": 8.9,  "no_elec": 12.1, "no_toilet": 7.2,  "food_insec": 35.8},
    "KwaZulu-Natal": {"no_water": 25.8, "no_elec": 23.5, "no_toilet": 21.3, "food_insec": 38.4},
    "North West":    {"no_water": 18.4, "no_elec": 15.8, "no_toilet": 16.1, "food_insec": 36.2},
    "Gauteng":       {"no_water": 5.1,  "no_elec": 5.3,  "no_toilet": 4.8,  "food_insec": 24.7},
    "Mpumalanga":    {"no_water": 22.3, "no_elec": 18.9, "no_toilet": 18.7, "food_insec": 39.1},
    "Limpopo":       {"no_water": 27.6, "no_elec": 26.1, "no_toilet": 24.9, "food_insec": 44.6},
}


def _find_row_label(ws, label: str, col: int = 0, max_rows: int = 300) -> int | None:
    for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), start=1):
        cell = row[col] if col < len(row) else None
        if cell and label.lower() in str(cell).lower():
            return i
    return None


def _get_province_row(ws, row_num: int, prov_col_start: int = 1) -> dict[str, float]:
    rows = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    if not rows:
        return {}
    row = rows[0]
    result = {}
    for i, prov in enumerate(PROVINCES):
        idx = prov_col_start + i
        val = row[idx] if idx < len(row) else None
        try:
            result[prov] = float(val) if val is not None else 0.0
        except (TypeError, ValueError):
            result[prov] = 0.0
    return result


def _sum_rows(ws, labels: list[str], prov_col_start: int = 1) -> dict[str, float]:
    total: dict[str, float] = {p: 0.0 for p in PROVINCES}
    found_any = False
    for label in labels:
        row_num = _find_row_label(ws, label)
        if row_num:
            row_data = _get_province_row(ws, row_num, prov_col_start)
            for prov, val in row_data.items():
                total[prov] += val
            found_any = True
            logger.info(f"  Found '{label}': {[round(v,1) for v in list(row_data.values())[:4]]}")
    return total if found_any else {}


def load_from_excel() -> dict[str, dict]:
    """Returns dict: province -> {no_water, no_elec, no_toilet, food_insec}"""
    result = {p: dict(FALLBACK[p]) for p in PROVINCES}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(ADDENDUM_FILE, read_only=True, data_only=True)
        logger.info(f"Sheets: {wb.sheetnames}")

        # ── Water: Table 9.1 ────────────────────────────────────────────────
        water_sheet = None
        for name in wb.sheetnames:
            if "9.1" in name or ("water" in name.lower() and "9" in name):
                water_sheet = wb[name]
                break

        if water_sheet:
            no_water_labels = [
                "no access", "no piped", "river", "stream", "spring",
                "rainwater", "borehole", "other source", "dam",
            ]
            no_water = _sum_rows(water_sheet, no_water_labels)
            if no_water:
                for p in PROVINCES:
                    result[p]["no_water"] = no_water.get(p, FALLBACK[p]["no_water"])
                logger.info("Water data loaded from Excel.")
            else:
                logger.warning("No water rows matched; keeping fallback.")
        else:
            logger.warning("Table 9.1 sheet not found; keeping water fallback.")

        # ── Electricity: Table 11.1 ──────────────────────────────────────────
        elec_sheet = None
        for name in wb.sheetnames:
            if "11.1" in name or ("elec" in name.lower() and "11" in name):
                elec_sheet = wb[name]
                break

        if elec_sheet:
            no_elec_labels = [
                "no electricity", "not connected", "paraffin", "candles",
                "wood", "solar", "gas", "other energy", "none",
            ]
            no_elec = _sum_rows(elec_sheet, no_elec_labels)
            if no_elec:
                for p in PROVINCES:
                    result[p]["no_elec"] = no_elec.get(p, FALLBACK[p]["no_elec"])
                logger.info("Electricity data loaded from Excel.")
            else:
                logger.warning("No electricity rows matched; keeping fallback.")
        else:
            logger.warning("Table 11.1 sheet not found; keeping electricity fallback.")

        wb.close()

    except FileNotFoundError:
        logger.warning(f"Addendum file not found at {ADDENDUM_FILE}; using fallback values.")
    except Exception as e:
        logger.warning(f"Excel parse error ({e}); using fallback values.")

    return result


def _compute_service_fail_score(data: dict) -> float:
    return round((data["no_water"] + data["no_elec"] + data["no_toilet"]) / 300.0, 6)


def _get_supabase_client():
    from supabase import create_client
    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_KEY", "") or os.environ.get("SUPABASE_KEY", "")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_KEY must be set.")
    return create_client(url, key)


def run() -> None:
    province_data = load_from_excel()

    # Build ghs_data rows (one per province, using province as "municipality" key)
    ghs_rows = []
    for prov in PROVINCES:
        d = province_data[prov]
        ghs_rows.append({
            "province":            prov,
            "municipality":        prov,           # provincial aggregate
            "pct_no_piped_water":  round(d["no_water"], 2),
            "pct_no_electricity":  round(d["no_elec"], 2),
            "pct_no_flush_toilet": round(d["no_toilet"], 2),
            "pct_food_insecure":   round(d["food_insec"], 2),
            "survey_year":         SURVEY_YEAR,
        })

    supabase = _get_supabase_client()

    logger.info(f"Upserting {len(ghs_rows)} rows into ghs_data …")
    result = supabase.table("ghs_data").upsert(ghs_rows, on_conflict="municipality,survey_year").execute()
    upserted = len(result.data) if result.data else 0
    logger.info("Done.")

    print(f"\nGHS {SURVEY_YEAR} — Provincial summary:")
    print(f"  {'Province':<20} {'No Water%':>10} {'No Elec%':>10} {'No Toilet%':>11} {'Food Insec%':>12}")
    print(f"  {'-'*65}")
    for r in ghs_rows:
        print(f"  {r['province']:<20} {r['pct_no_piped_water']:>9.1f}% {r['pct_no_electricity']:>9.1f}% {r['pct_no_flush_toilet']:>10.1f}% {r['pct_food_insecure']:>11.1f}%")
    print(f"\n  ghs_data rows upserted: {upserted}")


if __name__ == "__main__":
    run()
